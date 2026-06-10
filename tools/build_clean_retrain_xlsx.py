#!/usr/bin/env python
"""Render the clean-retrain results into docs/results/clean_retrain_summary.xlsx
— the single human-facing report (Chinese headers, color-grouped by experiment
family, glossary sheet). This .xlsx is the source of truth in docs/results/.

Row data comes straight from build_rows() in tools/build_clean_retrain_summary.py
(which reads the eval JSONs / scalars live) — no CSV middleman, so new experiments
appear automatically and in-flight rows render as "进行中". The old
docs/results/clean_retrain_summary_table.csv has been retired.

Run (no torch / GPU needed):
    python tools/build_clean_retrain_xlsx.py
"""
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from build_clean_retrain_summary import build_rows

ROOT = Path(__file__).resolve().parent.parent
XLSX_OUT = ROOT / 'docs' / 'results' / 'clean_retrain_summary.xlsx'

# ---------------------------------------------------------------------------
# Mapping tables (the only place readable Chinese lives — easy to edit).
# ---------------------------------------------------------------------------

# 实验家族顺序（行按此排序，使同色块连续）。文本侧家族放一起(回答"文本起不起作用")，
# 骨干放最后(回答"换骨干行不行")。
FAMILY_ORDER = ['基线', 'ICF·真实属性', '形态学属性', 'ICF·随机属性', '骨干']
FAMILY_FILL = {
    '基线':        'FFE8F0FE',  # 浅蓝
    '骨干':        'FFEDE7F6',  # 浅紫
    'ICF·真实属性': 'FFE6F4EA',  # 浅绿
    '形态学属性':   'FFFEF7E0',  # 浅黄
    'ICF·随机属性': 'FFFCE8E6',  # 浅红
}

# 原代号(Tag) -> (实验名称, 家族)
EXP = {
    'Row1':          ('基线 M1 · BiomedCLIP（当前最佳 base）', '基线'),
    'B1':            ('基线 M1 · XLM-R', '基线'),
    'dinov3_frozen_bmc': ('DINOv3-ConvNeXt冻结 · 单提示词 · BiomedCLIP', '骨干'),
    'dinov3_frozen_xlmr': ('DINOv3-ConvNeXt冻结 · 单提示词 · XLM-R', '骨干'),
    'dinov3vits_bmc': ('DINOv3 ViT-S/16冻结 · 单提示词 · BiomedCLIP', '骨干'),
    'dinov3vits_unfz8_ms_bmc': ('DINOv3 ViT-S/16部分微调(冻8训4)+多尺度 · BiomedCLIP', '骨干'),
    'dinov3cnx_fullft_bmc': ('DINOv3-ConvNeXt全量微调 · 单提示词 · BiomedCLIP', '骨干'),
    'A4':            ('ICF · 单提示词 · BiomedCLIP', 'ICF·真实属性'),
    'icf_5attr':     ('ICF · 真实5属性 · BiomedCLIP', 'ICF·真实属性'),
    'morph6_real':   ('ICF · 形态学6属性(真实) · BiomedCLIP', '形态学属性'),
    'morph6_mean':   ('形态学6属性 · 均值池化(无ICF) · BiomedCLIP', '形态学属性'),
    'morph6_random': ('ICF · 形态学6属性(随机) · BiomedCLIP', '形态学属性'),
    'Row17':         ('ICF · 随机5属性 · BiomedCLIP', 'ICF·随机属性'),
    'ICFrX':         ('ICF · 随机5属性 · XLM-R', 'ICF·随机属性'),
}

TEXTSRC = {
    '1-PSC real':               '单提示词/类（真实, 1-PSC）',
    'random':                   '随机属性',
    'real morphology':          '真实形态学属性（morph6）',
    '5-attr real(label-mixed)': '真实5属性（含标签混入）',
}

ICF_EXACT = {
    'degenerate':           '退化',
    'no-fusion':            '未融合（无 ICF）',
    'mean-pool(no ICF)':    '均值池化（无 ICF）',
    'no-fusion(DINOv3-bb)': '未融合（DINOv3 骨干）',
    'alive':                '融合生效',
}


def render_icf(v):
    if v in ICF_EXACT:
        return ICF_EXACT[v]
    m = re.match(r'^(dead|alive)\((cos[\d.]+)\)$', v)
    if m:
        head = '融合塌缩/失效' if m.group(1) == 'dead' else '融合生效'
        return '%s（%s）' % (head, m.group(2))
    return v  # 未知值原样保留，不让生成器崩


# ---------------------------------------------------------------------------
# 术语图例（Sheet2 与 docs/glossary.md 同源；改一处请同步另一处）
# ---------------------------------------------------------------------------
GLOSSARY = [
    ('ICF', 'Image-Conditional Fusion，图像条件文本融合'),
    ('M1', '不带融合的单提示词基线（1-prompt baseline）'),
    ('1-PSC', '单提示词/类（one prompt per class）'),
    ('morph6', '6 轴形态学属性文本'),
    ('5attr', '5 属性文本'),
    ('BiomedCLIP', '生物医学图文编码器（保留英文）'),
    ('XLM-R', '多语言文本编码器 XLM-RoBERTa（保留英文）'),
    ('DINOv3', '自监督视觉骨干（保留英文，在研）'),
    ('base25', '测试集中训练见过的 25 个基础类'),
    ('novel9', '9 个零样本新类（训练未见）'),
    ('mAP', 'mean Average Precision，平均精度（保留英文）'),
    ('base25 三口径', 'organ-macro=5器官等权（Serous 单类却占20%权重，会放大单类、误导）；class-macro=25类等权（更公平，判优以此为准）；instance-weighted=按实例数加权。novel9 类均衡，organ-macro=class-macro。'),
    ('随机种子', '全实验线训练此前未设固定种子→run-to-run 不可复现，≤2pp 的变体差异可能纯属噪声，需多 seed（≥3）取均值±std。default_runtime 现已加 seed=0。'),
    ('ICF 状态·退化', 'degenerate：ICF 退化'),
    ('ICF 状态·未融合', 'no-fusion：无 ICF 融合'),
    ('ICF 状态·均值池化', 'mean-pool：多属性直接取平均，未走 ICF'),
    ('ICF 状态·融合塌缩', 'dead(cos1.0)：融合失效，同类跨图相似度=1.0'),
    ('ICF 状态·融合生效', 'alive(cosX)：融合生效，同类跨图相似度 X<1'),
    ('进行中', '该实验评测/数据尚未产出（pending），待 paper_eval 完成后重跑刷新'),
    ('— 原代号 —', '——以下为汇总表「原代号」列的含义——'),
    ('Row1', '基线 M1 · BiomedCLIP（当前最佳 base）'),
    ('B1', '基线 M1 · XLM-R'),
    ('dinov3', 'DINOv3 骨干 · 单提示词 · BiomedCLIP（注：表中旧 run 因 SkipImageBackboneInLoadFromHook 挂错生命周期而失效，实为 ImageNet；hook 已于 2026-05-30 修复，需重跑才是真 DINOv3）'),
    ('A4', 'ICF · 单提示词 · BiomedCLIP'),
    ('icf_5attr', 'ICF · 真实5属性 · BiomedCLIP'),
    ('morph6_real', 'ICF · 形态学6属性(真实) · BiomedCLIP'),
    ('morph6_mean', '形态学6属性 · 均值池化(无ICF) · BiomedCLIP'),
    ('morph6_random', 'ICF · 形态学6属性(随机) · BiomedCLIP'),
    ('Row17', 'ICF · 随机5属性 · BiomedCLIP'),
    ('ICFrX', 'ICF · 随机5属性 · XLM-R'),
]

# 列定义：(中文表头, CSV 字段, 类型)  type ∈ {text, int, float, name}
# 高亮列(best 标金色)
BEST_FIELDS = ('base25_organ_macro', 'base25_class_macro', 'novel9_organ_macro')
# 精简版:只保留「方法 + 结果」。方法(编码器/真随机文本/属性数)已全部写进「方法」名,
# 故去掉 ICF状态 + 文本属性数/来源/编码器 三个冗余列。
COLUMNS = [
    ('方法', 'Tag', 'name'),
    ('base25\n器官宏平均\n(headline·会放大单类)', 'base25_organ_macro', 'float'),
    ('base25\n类宏平均\n(更公平·以此判优)', 'base25_class_macro', 'float'),
    ('base25\n实例加权', 'base25_inst_weighted', 'float'),
    ('novel9\n宏平均(零样本)', 'novel9_organ_macro', 'float'),
    ('novel9\n实例加权', 'novel9_inst_weighted', 'float'),
    ('验证集 mAP', 'val_mAP_macro', 'float'),
    ('最佳轮次', 'best_epoch', 'int'),
]
COL_WIDTH = [48, 13, 13, 11, 13, 11, 11, 8]

HEADER_FILL = PatternFill(start_color='FFD9D9D9', end_color='FFD9D9D9', fill_type='solid')
BEST_FILL = PatternFill(start_color='FFFFF2CC', end_color='FFFFF2CC', fill_type='solid')
THIN = Side(style='thin', color='FFBFBFBF')
GRID = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MEDIUM = Side(style='medium', color='FFBF9000')
BEST_BORDER = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)


def to_float(s):
    try:
        return float(s)
    except (TypeError, ValueError):
        return None


def best_of(rows, field):
    vals = [v for v in (to_float(r.get(field)) for r in rows) if v is not None]
    return max(vals) if vals else None


def render_table(ws, rows, note_text):
    """Render one results table (header + rows + per-table best-highlight + footnote).
    `best` is computed WITHIN `rows`, so each (sub-)table highlights its own best —
    e.g. the 文本消融 sub-table highlights the best among the text variants only."""
    headers = [c[0] for c in COLUMNS]
    ws.append(headers)
    for ci, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=ci)
        c.fill = HEADER_FILL
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = GRID
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 46

    best = {f: best_of(rows, f) for f in BEST_FIELDS}

    for ri, r in enumerate(rows, start=2):
        tag = r['Tag']
        name, fam = EXP.get(tag, (tag, '其它'))
        in_flight = to_float(r.get('base25_organ_macro')) is None
        if in_flight:
            name = name + '（评测进行中）'
        fill_argb = FAMILY_FILL.get(fam, 'FFFFFFFF')
        row_fill = PatternFill(start_color=fill_argb, end_color=fill_argb, fill_type='solid')

        for ci, (_, field, typ) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=ri, column=ci)
            cell.border = GRID
            cell.fill = row_fill
            raw_v = r.get(field, '')

            if typ == 'name':
                cell.value = name
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                cell.font = Font(bold=True)
                continue
            if typ == 'textsrc':
                cell.value = TEXTSRC.get(raw_v, raw_v)
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                continue
            if typ == 'icf':
                cell.value = render_icf(raw_v)
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                continue
            if typ == 'text':
                cell.value = raw_v
                cell.alignment = Alignment(horizontal='center', vertical='center')
                continue

            fv = to_float(raw_v)
            if fv is None:
                cell.value = '进行中'
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(color='FF999999', italic=True)
                continue
            if typ == 'int':
                cell.value = int(fv)
            else:
                cell.value = fv
                cell.number_format = '0.0000'
            cell.alignment = Alignment(horizontal='center', vertical='center')

            if field in best and best[field] is not None and fv == best[field]:
                cell.fill = BEST_FILL
                cell.font = Font(bold=True)
                cell.border = BEST_BORDER

    nr = len(rows) + 2
    ws.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=len(COLUMNS))
    note = ws.cell(row=nr, column=1)
    note.value = note_text
    note.alignment = Alignment(wrap_text=True, vertical='top')
    note.font = Font(italic=True, color='FF7F6000')
    ws.row_dimensions[nr].height = 72

    for ci, w in enumerate(COL_WIDTH, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w


MAIN_NOTE = (
    '注：base25「器官宏平均」是历史 headline，但 Serous 器官仅 1 个类却占 1/5 权重，会放大单类——'
    '换「类宏平均」后各变体接近平局，Row1 反而偏低，故判优以类宏平均为准。'
    '全实验线此前无随机种子，≤2pp 差异可能是噪声，需多 seed 确认。'
    '按分析维度看请翻「①文本消融」「②Backbone消融」子表。完整问题分析见 docs/results/problem_analysis_20260530.md。')

TEXT_NOTE = (
    '【文本消融——回答"文本到底起不起作用"】同一 ImageNet-ConvNeXt 骨干，只变文本侧(真实/随机 × 5属性/6属性 × ICF/均值池化)。'
    '读法(类宏)：真实5属性 0.3242 vs 随机5属性 0.3182 = +0.6pp 且 19/25 类胜随机(sign-test p≈0.007)'
    '→ 含类名的临床文本(诊断名+IHC)有【小但显著】的边际，不是纯噪声；'
    '而 label-free morph6 真实 0.3181 ≈ 随机 0.3253 ≈ M1基线 0.3177 → 纯形态学文本才是惰性的。'
    'magnitude 被 logit_scale≈0.368 压小；待多 seed + 结构-vs-类名 消融钉死(见 feedback_text_branch_inert_for_base 修正版)。')

BB_NOTE = (
    '【Backbone 消融——回答"换图像骨干行不行"】全部 1-PSC 单提示词文本，只换骨干。'
    'Row1=ImageNet/COCO-ConvNeXt(端到端训练，类宏 0.3177，最佳)。DINOv3 各变体(冻结/全量微调/ViT-S/多尺度)'
    '类宏 0.195–0.299 全部 < Row1。→ 自然图自监督骨干不适合密集细胞学检测；'
    'A(DINOv3-ConvNeXt 全微调 0.2993)是干净单变量对照，确认 DINOv3 初始化对 ConvNeXt 检测确实更差。')


def main():
    raw = build_rows()
    by_tag = {r['Tag']: r for r in raw}

    # main sheet: all rows, sorted by family (text families together, 骨干 last)
    def fam_key(r):
        fam = EXP.get(r['Tag'], (r['Tag'], '其它'))[1]
        return (FAMILY_ORDER.index(fam) if fam in FAMILY_ORDER else len(FAMILY_ORDER),)
    rows_all = sorted(raw, key=fam_key)

    wb = Workbook()
    ws = wb.active
    ws.title = '结果汇总'
    render_table(ws, rows_all, MAIN_NOTE)

    # analysis sub-tables: cross-family, curated order per dimension
    TEXT_TAGS = ['Row1', 'B1', 'A4', 'icf_5attr', 'morph6_real', 'morph6_mean',
                 'morph6_random', 'Row17', 'ICFrX']
    BB_TAGS = ['Row1', 'dinov3_frozen_bmc', 'dinov3_frozen_xlmr',
               'dinov3cnx_fullft_bmc', 'dinov3vits_bmc', 'dinov3vits_unfz8_ms_bmc']
    text_rows = [by_tag[t] for t in TEXT_TAGS if t in by_tag]
    bb_rows = [by_tag[t] for t in BB_TAGS if t in by_tag]
    render_table(wb.create_sheet('①文本消融'), text_rows, TEXT_NOTE)
    render_table(wb.create_sheet('②Backbone消融'), bb_rows, BB_NOTE)

    # ---- Sheet: 术语图例 ----
    ws2 = wb.create_sheet('术语图例')
    ws2.append(['术语', '含义'])
    for ci in (1, 2):
        c = ws2.cell(row=1, column=ci)
        c.fill = HEADER_FILL
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = GRID
    ws2.freeze_panes = 'A2'
    for ri, (term, meaning) in enumerate(GLOSSARY, start=2):
        a = ws2.cell(row=ri, column=1, value=term)
        b = ws2.cell(row=ri, column=2, value=meaning)
        a.font = Font(bold=True)
        a.alignment = Alignment(vertical='center')
        b.alignment = Alignment(vertical='center', wrap_text=True)
        a.border = GRID
        b.border = GRID
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 62

    wb.save(XLSX_OUT)
    print('wrote %s  (%d experiments; 子表 文本%d / Backbone%d; %d glossary terms)' % (
        XLSX_OUT, len(rows_all), len(text_rows), len(bb_rows), len(GLOSSARY)))


if __name__ == '__main__':
    main()
